import math
import datetime

import openpyxl as op
import pandas as pd
from openpyxl.styles import Border, Side, Font
from openpyxl.utils import get_column_letter


vacant_dic = {"name": "Название", "description": "Описание", "key_skills": "Навыки", "experience_id": "Опыт работы",
              "premium": "Премиум-вакансия", "employer_name": "Компания", "salary": "Оклад",
              "area_name": "Название региона", "published_at": "Дата публикации вакансии"}
currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
curr_dict = {}

class InputConect():
    dataSet = ""
    dict_inYear_noName = {}
    dict_inYear_noName_salary = {}

    dict_inYear_WithName = {}
    dict_inYear_WithName_salary = {}

    dict_inYear_City = {}
    dict_inYear_City_salary = {}

    skill_dict = {}
    temp_dict = {}
    temp_salary_dict = {}

    def __init__(self, data, skip_init = False):
        if not skip_init:
            self.dataSet = data

    def find_full_skills(self, x, full_skills):
        """
        Правило сортировки по всем скиллам. Осуществляет поиск полного списка навыков для каждой строчки (не укороченного)

        :param x (list) - список элементов одной вакансии
        :param full_skills (dict) - словарь полных навыков для каждой профессии
        :return: string содержащий все навыки для вакансии хранящейся в строке x
        """
        list_skills = full_skills[x[1]]
        redact_skills = x[3].replace(".", "")
        index = 0
        for i in range(len(list_skills)):
            str_skills = '\n'.join(list_skills[i])
            if str(str_skills).__contains__(redact_skills):
                index = i
                break
        return list_skills[index]

    def get_date_sort(self, x):
        city = x.area_name
        year = x.published_at[0:4]

        try:
            f = currency_to_rub[x.salary.salary_currency]
        except:
            currency_to_rub[x.salary.salary_currency] = 0

        if not self.dict_inYear_WithName.__contains__(int(year)) and (x.name.__contains__(dataSet.job_name)
                                                                            or x.name.__contains__(
                    dataSet.job_name.lower())):
            self.dict_inYear_WithName[int(year)] = 1
            self.dict_inYear_WithName_salary[int(year)] = (float(x.salary.salary_from)
                                                                 * currency_to_rub[x.salary.salary_currency]
                                                                 + float(x.salary.salary_to)
                                                                 * currency_to_rub[x.salary.salary_currency]) \
                                                                / 2
            if len(x.key_skills) > 0:
                for skill in x.key_skills[0].split('\n'):
                    if not self.skill_dict.__contains__(skill):
                        self.skill_dict[skill] = 1
                    elif self.skill_dict.__contains__(skill):
                        self.skill_dict[skill] += 1
        elif x.name.__contains__(dataSet.job_name) or x.name.__contains__(dataSet.job_name.lower()):
            self.dict_inYear_WithName[int(year)] += 1
            self.dict_inYear_WithName_salary[int(year)] += (float(x.salary.salary_from)
                                                                  * currency_to_rub[x.salary.salary_currency]
                                                                  + float(x.salary.salary_to)
                                                                  * currency_to_rub[x.salary.salary_currency]) \
                                                                 / 2
            if len(x.key_skills) > 0:
                for skill in x.key_skills[0].split('\n'):
                    if not self.skill_dict.__contains__(skill):
                        self.skill_dict[skill] = 1
                    elif self.skill_dict.__contains__(skill):
                        self.skill_dict[skill] += 1
        if not self.dict_inYear_noName.__contains__(int(year)):
            self.dict_inYear_noName[int(year)] = 1
            self.dict_inYear_noName_salary[int(year)] = (float(x.salary.salary_from)
                                                               * currency_to_rub[x.salary.salary_currency]
                                                               + float(x.salary.salary_to)
                                                               * currency_to_rub[x.salary.salary_currency]) \
                                                              / 2
        elif self.dict_inYear_noName.__contains__(int(year)):
            self.dict_inYear_noName[int(year)] += 1
            self.dict_inYear_noName_salary[int(year)] += (float(x.salary.salary_from)
                                                                * currency_to_rub[x.salary.salary_currency]
                                                                + float(x.salary.salary_to)
                                                                * currency_to_rub[x.salary.salary_currency]) \
                                                               / 2
        if not self.temp_dict.__contains__(city):
            self.temp_dict[city] = 1
            self.temp_salary_dict[city] = (float(x.salary.salary_from)
                                              * currency_to_rub[x.salary.salary_currency]
                                              + float(x.salary.salary_to)
                                              * currency_to_rub[x.salary.salary_currency]) \
                                             / 2
        elif self.temp_dict.__contains__(city):
            self.temp_dict[city] += 1
            self.temp_salary_dict[city] += (float(x.salary.salary_from)
                                               * currency_to_rub[x.salary.salary_currency]
                                               + float(x.salary.salary_to)
                                               * currency_to_rub[x.salary.salary_currency]) \
                                              / 2

        return datetime.datetime(day=int(x.published_at[8:10]),
                                 month=int(x.published_at[5:7]),
                                 year=int(x.published_at[0:4]),
                                 hour=int(x.published_at[11:13]),
                                 minute=int(x.published_at[14:16]),
                                 second=int(x.published_at[17:19]))

    def get_year_sort(self, x):
        split_list = x.split(" ")
        max_year = 0
        count_ind = 0
        for word in split_list:
            if word.isdigit():
                count_ind += 1
                max_year = max(max_year, int(word))
        if count_ind == 1:
            max_year += 1
        return max_year

    def get_sort_table(self, dataSet, table, full_skills):
        column_for_sort = dataSet.sort_parameter
        index_column = list(table.field_names).index(column_for_sort)
        reversesort = dataSet.IsReverseSort
        if column_for_sort == "Навыки":
            sort_key = lambda x: len(self.find_full_skills(x, full_skills))
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort == "Дата публикации вакансии":
            sort_key = lambda x: self.get_date_sort(x)
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort == "Оклад":
            salar = Salary()
            sort_key = lambda x: salar.salary_sorter(x[index_column])
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort == "Опыт работы":
            sort_key = lambda x: self.get_year_sort(x[index_column])
            return sorted(table.rows, key=sort_key, reverse=reversesort)
        elif column_for_sort != "":
            sort_key = lambda x: x[index_column]
            return sorted(table.rows, key=sort_key, reverse=reversesort)

    def get_sort_dataSet(self, dataSet):
        reversesort = dataSet.IsReverseSort
        sort_key = lambda x: self.get_date_sort(x)
        return sorted(dataSet.vacancies_objects, key=sort_key, reverse=reversesort)

    def sorted_for_graf(self):
        dataSet.sort_parameter = "Дата публикации вакансии"
        self.dataSet = self.get_sort_dataSet(self.dataSet)
        for key in self.dict_inYear_noName.keys():
            self.dict_inYear_noName_salary[key] = math.floor(
                int(self.dict_inYear_noName_salary[key]) / self.dict_inYear_noName[key])
            if len(self.dict_inYear_WithName) > 0:
                self.dict_inYear_WithName_salary[key] = math.floor(
                    int(self.dict_inYear_WithName_salary[key]) / self.dict_inYear_WithName[key])
            else:
                self.dict_inYear_WithName[key] = 0
                self.dict_inYear_WithName_salary[key] = 0
        bad_city_vac_count = 0
        for city in self.temp_dict.keys():
            if self.temp_dict[city] >= math.floor(len(dataSet.vacancies_objects) / 100):
                try:
                    self.dict_inYear_City[city] = round(int(self.temp_dict[city]) / len(dataSet.vacancies_objects), 4)
                    self.dict_inYear_City_salary[city] = math.floor(
                        int(self.temp_salary_dict[city]) / self.temp_dict[city])
                except:
                    f = 6
            else:
                bad_city_vac_count += int(self.temp_dict[city]) / len(dataSet.vacancies_objects)
        self.dict_inYear_City["Другие"] = bad_city_vac_count


class DataSet():
    file_name = ""
    job_name = ""
    city_name = ""

    vacancies_objects = []
    title_piece = ["№"] + list(vacant_dic.values())
    sort_parameter = ""
    IsReverseSort = False
    filter_for_table = []
    vacant_piece = []

    message_error = ""
    filter_atr = ""
    sort_atr = ""
    revers_atr = ""

    def __init__(self):
        self.file_name = "csv1/vacancies_with_skills.csv"
        self.job_name = "Инженер-программист"
        self.vacancies_objects = self.csv_filter_pandas(self.file_name)
        if len(self.vacancies_objects) == 0:
            print("Нет данных")
            exit()


    def csv_filter_pandas(self,file):
        df = pd.read_csv(file, encoding='utf_8_sig')
        vacant_list = []
        for row in df.iterrows():
            vacant_list.append(Vacancy(row[1]))
        return vacant_list


class Vacancy():
    name = ""
    description = ""
    key_skills = []
    experience_id = ""
    premium = ""
    employer_name = ""
    salary = ""
    area_name = ""
    published_at = ""

    def __init__(self, vacant):
        self.name = vacant["name"]
        gross = "None"
        try:
            self.key_skills = vacant["key_skills"].split(']')
            self.experience_id = vacant["experience_id"]
            self.premium = vacant["premium"]
            self.employer_name = vacant["employer_name"]
            gross = vacant["salary_gross"]
            self.description = vacant["description"]
        except:
            f = 5
        self.salary = Salary(vacant["salary_from"], vacant["salary_to"], gross,
                             vacant["salary_currency"])
        self.area_name = vacant["area_name"]
        self.published_at = vacant["published_at"]


class Salary():
    salary_from = ""
    salary_to = ""
    salary_gross = ""
    salary_currency = ""

    def __init__(self, *args):
        global curr_dict
        if len(args) > 0:
            self.salary_from = self.ternik(args[0])
            self.salary_to = self.ternik(args[1])
            if args[2] != "None":
                self.salary_gross = self.ternik(args[2])
            self.salary_currency = self.ternik(args[3])
            if not curr_dict.__contains__(args[3]):
                curr_dict[args[3]] = 1
            else:
                curr_dict[args[3]] += 1

    def ternik(self, obj):
        if pd.isna(obj):
            return 0
        else:
            return obj


    def prepare_salary(self, string_salary):
            list_numb = []
            count = 0
            for char in reversed(string_salary):
                if count < 3:
                    list_numb.append(char)
                    count += 1
                else:
                    list_numb.append(" ")
                    list_numb.append(char)
                    count = 0
            return "".join(list_numb.__reversed__())

    def salary_sorter(self, x):
        currency = x.salary_currency
        salar_min = float(x.salary_from) * currency_to_rub[currency]
        salar_max = float(x.salary_to) * currency_to_rub[currency]
        return (salar_min + salar_max) / 2


class report():
    border = 0
    font = 0

    total_year = []
    mean_salary = {}
    mean_salary_job = {}
    count_vac = {}
    count_vac_job = {}

    mean_salary_city = {}
    count_vac_city = {}

    skills = {}

    book = 0

    def __init__(self, font, border):
        self.font = font
        self.border = border

    def generate_excel(self, years, data, book, name_job):

        self.total_year = years
        self.mean_salary = data[0]
        self.mean_salary_job = data[1]
        self.count_vac = data[2]
        self.count_vac_job = data[3]
        self.mean_salary_city = data[4]
        self.count_vac_city = data[5]
        self.skills = data[6]

        # 1
        ws = book.active
        ws.title = "Статистика по годам"
        book.create_sheet("Статистика по городам", 1)
        book.create_sheet("Топ-10 навыков", 2)
        ws['A1'] = "Год"
        ws['A1'].font = self.font

        ws['B1'] = "Средняя зарплата"
        ws['B1'].font = self.font

        ws['C1'] = "Средняя зарплата - {0}".format(name_job)
        ws['C1'].font = self.font

        ws['D1'] = "Количество вакансий"
        ws['D1'].font = self.font

        ws['E1'] = "Количество вакансий - {0}".format(name_job)
        ws['E1'].font = self.font

        for i in range(len(years)):
            ws['A{0}'.format(i + 2)] = years[i]
            ws['B{0}'.format(i + 2)] = self.mean_salary[years[i]]
            ws['C{0}'.format(i + 2)] = self.mean_salary_job[years[i]]
            ws['D{0}'.format(i + 2)] = self.count_vac[years[i]]
            ws['E{0}'.format(i + 2)] = self.count_vac_job[years[i]]

        column_widths = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for row in ws.rows:
            for i, cell in enumerate(row):
                column_widths[i + 1] = max(len((str)(cell.value)) + 1, column_widths[i + 1])
                cell.border = self.border
        for i in range(len(column_widths)):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i + 1]

        # 2
        ws = book["Статистика по городам"]
        ws['A1'] = "Город"
        ws['A1'].font = self.font

        ws['B1'] = "Уровень зарплат"
        ws['B1'].font = self.font

        ws['D1'] = "Город"
        ws['D1'].font = self.font

        ws['E1'] = "Доля вакансий".format(name_job)
        ws['E1'].font = self.font

        cityes_salar = list(self.mean_salary_city.keys())
        cityes_vac = list(self.count_vac_city.keys())
        for i in range(len(cityes_salar)):
            ws['A{0}'.format(i + 2)] = cityes_salar[i]
            ws['B{0}'.format(i + 2)] = self.mean_salary_city[cityes_salar[i]]
            ws['D{0}'.format(i + 2)] = cityes_vac[i]
            ws['E{0}'.format(i + 2)] = self.count_vac_city[cityes_vac[i]]
            ws['E{0}'.format(i + 2)].number_format = "0%"

        column_widths = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        for row in ws.rows:
            for i, cell in enumerate(row):
                column_widths[i + 1] = max(len((str)(cell.value)) + 1, column_widths[i + 1])
                cell.border = self.border
        for i in range(len(column_widths)):
            ws.column_dimensions[get_column_letter(i + 1)].width = column_widths[i + 1]
        # 3
        ws = book["Топ-10 навыков"]

        skill_list = list(self.skills.keys())
        skill_count_list = list(self.skills.values())
        for i in range(len(skill_list) - 1):
            letter = chr(ord('A') + i)
            ws['{0}1'.format(letter)] = "{0}".format(skill_list[i])
            ws['{0}1'.format(letter)].font = self.font
            ws['{0}2'.format(letter)] = "{0}".format(skill_count_list[i])
            ws['{0}2'.format(letter)].font = self.font

        column_widths = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0,6:0,7:0,8:0,9:0,10:0}
        for row in ws.rows:
            for n, cell in enumerate(row):
                column_widths[n + 1] = max(len((str)(cell.value)) + 1, column_widths[n + 1])
                cell.border = self.border
        for n in range(len(column_widths)):
            ws.column_dimensions[get_column_letter(n + 1)].width = column_widths[n + 1]


dataSet = DataSet()
full_table_skills = {}
full_table_date = {}


# Динамика зарплат по годам
sorter_master = InputConect(dataSet)
# Сложная функция
sorter_master.sorted_for_graf()

sorter_master.dict_inYear_City_salary = dict(
    sorted(sorter_master.dict_inYear_City_salary.items(), key=lambda item: item[1], reverse=True))
sorter_master.dict_inYear_City = dict(
    sorted(sorter_master.dict_inYear_City.items(), key=lambda item: item[1], reverse=True))
sorter_master.skill_dict = dict(
    sorted(sorter_master.skill_dict.items(), key=lambda item: item[1], reverse=True)[:10])

sumInList = sorter_master.dict_inYear_City["Другие"] + sum(list(dict(list(sorter_master.dict_inYear_City.items())[10:]).values()))
sorter_master.dict_inYear_City["Другие"] = 0
sorter_master.dict_inYear_City = dict(
    sorted(sorter_master.dict_inYear_City.items(), key=lambda item: item[1], reverse=True))
sorter_master.dict_inYear_City = dict(list(sorter_master.dict_inYear_City.items())[:10])
sorter_master.dict_inYear_City_salary = dict(list(sorter_master.dict_inYear_City_salary.items())[:10])

print("Динамика уровня зарплат по годам: {0}".format(sorter_master.dict_inYear_noName_salary))
print("Динамика количества вакансий по годам: {0}".format(sorter_master.dict_inYear_noName))
print("Динамика уровня зарплат по годам для выбранной профессии: {0}".format(sorter_master.dict_inYear_WithName_salary))
print("Динамика количества вакансий по годам для выбранной профессии: {0}".format(sorter_master.dict_inYear_WithName))
print("Топ 10 навыков для профессии{0}: {1}".format(sorter_master.dict_inYear_WithName, sorter_master.skill_dict))
print("Уровень зарплат по городам (в порядке убывания): {0}".format(sorter_master.dict_inYear_City_salary))
print("Доля вакансий по городам (в порядке убывания): {0}".format(sorter_master.dict_inYear_City))

font_title = Font(name='Calibri',
                  size=11,
                  bold=True,
                  italic=False,
                  vertAlign=None,
                  underline='none',
                  strike=False,
                  color='FF000000')
border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))

wb = op.Workbook()

#Таблица
rep = report(font_title,border)
data_for_excel = [sorter_master.dict_inYear_noName_salary,
           sorter_master.dict_inYear_WithName_salary,
            sorter_master.dict_inYear_noName,
            sorter_master.dict_inYear_WithName,
            sorter_master.dict_inYear_City_salary,
            sorter_master.dict_inYear_City,
                  sorter_master.skill_dict]
rep.generate_excel(list(sorter_master.dict_inYear_noName_salary.keys()),data_for_excel,wb,dataSet.job_name)
wb.save("report.xlsx")

# #Графики
# labels_years = list(sorter_master.dict_inYear_noName_salary.keys())
# salary_noName = list(sorter_master.dict_inYear_noName_salary.values())
# salart_Name = list(sorter_master.dict_inYear_WithName_salary.values())
#
# vac_noName = list(sorter_master.dict_inYear_noName.values())
# vac_Name = list(sorter_master.dict_inYear_WithName.values())
#
# cityes_salary = list(sorter_master.dict_inYear_City_salary.values())
# labels_cityes = list(sorter_master.dict_inYear_City.keys())
#
# skills = list(sorter_master.skill_dict.values())
# labels_skills = list(sorter_master.skill_dict.keys())
#
# sorter_master.dict_inYear_City["Другие"] = sumInList
# sorter_master.dict_inYear_City = dict(
#     sorted(sorter_master.dict_inYear_City.items(), key=lambda item: item[1], reverse=True))
# circle_labels = list(sorter_master.dict_inYear_City.keys())
# cityes_perc = list(sorter_master.dict_inYear_City.values())
#
# width = 0.4
# x = np.arange(len(labels_years))
# y = np.arange(len(labels_cityes))
# y2 = np.arange(len(labels_skills))
#
# matplotlib.rc('axes', titlesize=8)
# matplotlib.rc('font', size=8)
# matplotlib.rc('xtick', labelsize=8)
# matplotlib.rc('ytick', labelsize=8)
# matplotlib.rc('legend', fontsize=8)
#
#
# # Для востребованности
#
# rects1 = plt.bar(x - width / 2, salary_noName, width, label="Средняя з/п", color='gold')
# plt.title('Уровень зарплат по годам')
# plt.xticks(x)
# plt.xlabel(labels_years, rotation=90)
# plt.legend()
# plt.savefig("graph1.png")
# plt.clf()
#
# rects2 = plt.bar(x + width / 2, salart_Name, width, label="з/п {0}".format(dataSet.job_name), color='orange')
# plt.title('Уровень зарплат по годам для профессии {0}'.format(dataSet.job_name))
# plt.xticks(x)
# plt.xlabel(labels_years, rotation=90)
# plt.legend()
# plt.savefig("graph2.png")
# plt.clf()
#
#
# rects3 = plt.bar(x - width / 2, vac_noName, width, label="Количество вакансий", color='green')
# plt.title('Количество вакансий по годам')
# plt.xticks(x)
# plt.xlabel(labels_years, rotation=90)
# plt.legend()
# plt.savefig("graph3.png")
# plt.clf()
#
# rects4 = plt.bar(x + width / 2, vac_Name, width, label="Количество вакансий {0}".format(dataSet.job_name), color='darkgreen')
# plt.title('Количество вакансий по годам для профессии {0}'.format(dataSet.job_name))
# plt.xticks(x)
# plt.xlabel(labels_years, rotation=90)
# plt.legend()
# plt.savefig("graph4.png")
# plt.clf()
#
# # Для географии
# rects5 = plt.barh(y, cityes_salary, width * 1.5, align='center')
# plt.title('Уровень зарплат по городам')
# plt.yticks(y, labels=labels_cityes)
# plt.gca().set_yticklabels(labels_cityes, fontsize=5.5,
#                          fontdict={'horizontalalignment': 'right', 'verticalalignment': 'center'})
# plt.gca().invert_yaxis()
# plt.savefig("graph5.png")
# plt.clf()
#
# circle = plt.pie(cityes_perc, labels=circle_labels, textprops={'fontsize': 6})
# plt.gca().set_title('Доля вакансий по городам', fontsize=6)
# plt.gca().axis('equal')
#
# plt.savefig("graph6.png")
# plt.clf()
#
#
# # Для навыков
# rects7 = plt.barh(y2, skills, width * 1.5, align='center')
# plt.title('Топ-10 навыков для профессии {0}'.format(dataSet.job_name))
# plt.yticks(y2, labels=labels_skills)
# plt.gca().set_yticklabels(labels_skills, fontsize=7,
#                          fontdict={'horizontalalignment': 'right', 'verticalalignment': 'center'})
# plt.gca().invert_yaxis()
# plt.savefig("graph7.png")
